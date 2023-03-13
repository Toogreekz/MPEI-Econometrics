import openpyxl
import math


def column(array):
    array.append(sum(array))
    array.append(sum(array[:len(array) - 1]) / (len(array) - 1))
    return array


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EconometricsTable"

N = int(input("Введи длину столбца х и у: "))
x = column([float(input()) for i in range(N)])
y = column([float(input()) for i in range(N)])
Xlog = column([math.log10(x[i]) for i in range(N)])
Ylog = column([math.log10(y[i]) for i in range(N)])
xy = column([x[i] * y[i] for i in range(N)])
x_2 = column([x[i] ** 2 for i in range(N)])
x_minus_x_sr = column([(x[i] - x[-1]) ** 2 for i in range(N)])
y_minus_y_sr = column([(y[i] - y[-1]) ** 2 for i in range(N)])
b = (xy[-1] - x[-1] * y[-1]) / (x_2[-1] - x[-1] * x[-1])
a = y[-1] - b * x[-1]
y_r = column([a + b * x[i] for i in range(N)])
y_minus_y_r = column([y[i] - y_r[i] for i in range(N)])
y_minus_y_r_2 = column([(y[i] - y_r[i]) ** 2 for i in range(N)])
A = column([abs(y_minus_y_r[i] / y[i]) * 100 for i in range(N)])

array = [x, y, Xlog, Ylog, xy, x_2, x_minus_x_sr, y_minus_y_sr, y_r, y_minus_y_r, y_minus_y_r_2, A]
row = ["x", "y", "X", "Y", "xy", "x^2", "(x-СРЕД(х))^2", "(y-СРЕД(y))^2", "ŷ", "y-ŷ",
         "(y-ŷ)^2", "A"]
s = "ABCDEFGHIJKL"
for i in s:
    ws[f"{i}1"] = row[s.index(i)]
for i in s:
    for elem in array[s.index(i)]:
        ws[f"{i}{array[s.index(i)].index(elem) + 2}"] = elem

sigmX = round(x_minus_x_sr[-1] ** 0.5, 5)
sigmY = round(y_minus_y_sr[-1] ** 0.5, 5)
sigmOST = round(y_minus_y_r_2[-1] ** 0.5, 5)
rXY = round(b * sigmX / sigmY, 5)
R = round((rXY ** 2) * 100, 5)
elast = round(b * x[-1] / y[-1], 5)
beta = rXY
F = round((R / (1 - R)) * (N - 2), 5)
ma = round(((sigmOST ** 2) * x_2[-2] / N ** 2 * sigmX ** 2) ** 0.5, 5)
mb = round((sigmOST ** 2 / N * sigmX ** 2) ** 0.5, 5)
mr = round(((1 - rXY ** 2) / (N - 2)) ** 0.5, 5)
try:
    ta = round(a / ma, 5)
except ZeroDivisionError:
    ta = "ZeroDivisionException: ma = 0"
try:
    tb = round(b / mb, 5)
except ZeroDivisionError:
    tb = "ZeroDivisionException: mb = 0"
try:
    tr = round(rXY / mr, 5)
except ZeroDivisionError:
    tr = "ZeroDivisionException: mr = 0"

ws["N1"] = f"sigmX = {sigmX}"
ws["N2"] = f"sigmY = {sigmY}"
ws["N3"] = f"sigmOST = {sigmOST}"
ws["N4"] = f"rXY = {rXY}"
ws["N5"] = f"R = {R}"
ws["N6"] = f"elast = {elast}"
ws["N7"] = f"beta = {beta}"
ws["N8"] = f"F = {F}"
ws["N9"] = f"ma = {ma}"
ws["N10"] = f"mb = {mb}"
ws["N11"] = f"mr = {mr}"
ws["N12"] = f"ta = {ta}"
ws["N13"] = f"tb = {tb}"
ws["N14"] = f"tr = {tr}"

if A[-1] <= 10:
    ws["N16"] = "Точная модель, среднее A <= 10"
else:
    ws["N16"] = "Неточная модель, среднее A > 10"
if rXY >= 0.7:
    ws["N17"] = "Связь сильная, rXY >= 0.7"
else:
    ws["N17"] = "Связь слабая, rXY < 0.7"
ws["N18"] = f"y зависит от х на {R}"
ws["N19"] = f"C увелечением x на 1%, у будет увеличиваться на {elast}"
ws["N20"] = f"С увелечением x на значение {sigmX} y будет увеличиваться на значение {beta}"

wb.save(filename="Econometrics.xlsx")
